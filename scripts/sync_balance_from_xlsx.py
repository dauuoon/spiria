from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - local setup issue
    raise SystemExit(
        "openpyxl is required. Install it with: python3 -m pip install openpyxl"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / 'balance' / 'Spiria_Game_Balance_v1.0.xlsx'


def read_sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [sheet.cell(2, column).value for column in range(1, sheet.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=True):
      if not any(value is not None and value != '' for value in row):
        continue
      item = {str(headers[index]).strip(): row[index] for index in range(len(headers)) if headers[index]}
      rows.append(item)
    return rows


def ts_string(value: Any) -> str:
    if value is None or value == '':
        return "''"
    if isinstance(value, bool):
        return 'true' if value else 'false'
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value)
    text = str(value)
    text = text.replace('\\', '\\\\').replace("'", "\\'")
    return f"'{text}'"


def format_int(value: int) -> str:
    return f'{value:,}'.replace(',', '_')


def write_if_changed(path: Path, text: str) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    current = path.read_text(encoding='utf-8') if path.exists() else None
    if current == text:
        return False
    path.write_text(text, encoding='utf-8')
    return True


def build_constants(player_rows: list[dict[str, Any]], quest_rows: list[dict[str, Any]], economy_rows: list[dict[str, Any]]) -> str:
    lookup = {row['키']: row['value'] for row in player_rows}
    quest_lookup = {row['키']: row['value'] for row in quest_rows}
    economy_lookup = {row['키']: row['value'] for row in economy_rows}
    return '\n'.join([
        '// Central constants for Spiria game runtime data',
        '// Source of truth: balance/Spiria_Game_Balance_v1.0.xlsx',
        '',
        f"export const MAX_LEVEL = {int(lookup['max_level'])} as const",
        f"export const TOTAL_EXP_TO_MAX = {format_int(int(lookup['total_exp_to_max']))} as const",
        '',
        f"export const INITIAL_MANA = {int(lookup['initial_mana'])} as const",
        f"export const MAX_MANA = {int(lookup['max_mana'])} as const",
        f"export const MANA_PER_EXPLORE = {int(lookup['mana_per_explore'])} as const",
        f"export const MANA_REGEN_MINUTES = {int(lookup['mana_regen_minutes'])} as const // 20 minutes per +1 mana",
        'export const MANA_REGEN_MS = MANA_REGEN_MINUTES * 60 * 1000',
        "export const MANA_STORAGE_KEY = 'spiria.mana' as const",
        "export const LEGACY_ENERGY_STORAGE_KEY = 'spiria.energy' as const",
        '',
        f"export const QUEST_BOARD_MAX_SLOTS = {int(quest_lookup['max_slots'])} as const",
        f"export const QUEST_RESPAWN_MINUTES = {int(quest_lookup['respawn_minutes'])} as const",
        '',
        f"export const MAIN_COLOR = {ts_string(economy_lookup['main_color'])} as const",
        '',
    ])


def build_levels(level_rows: list[dict[str, Any]]) -> str:
    numeric_rows = [row for row in level_rows if isinstance(row.get('from'), (int, float))]
    bands = [row for row in numeric_rows if int(row['from']) != 99]
    max_row = next(row for row in numeric_rows if int(row['from']) == 99)
    band_lines = []
    for row in bands:
        band_lines.append(
            f"  {{ from: {int(row['from'])}, to: {int(row['to'])}, exp: {int(row['exp_to_next(value)'])}, color: {ts_string(row['색상'])} }},"
        )
    return '\n'.join([
        "import type { LevelEntry, HexColor } from '../types/game'",
        "import { MAX_LEVEL } from './constants'",
        '',
        'const ranges: Array<{ from: number; to: number; exp: number; color: HexColor }> = [',
        *band_lines,
        ']',
        '',
        'const levels: LevelEntry[] = []',
        'for (const r of ranges) {',
        '  for (let lv = r.from; lv <= r.to; lv++) {',
        '    levels.push({ level: lv, expToNext: r.exp, color: r.color })',
        '  }',
        '}',
        f"levels.push({{ level: MAX_LEVEL, expToNext: {int(max_row['exp_to_next(value)'])}, color: {ts_string(max_row['색상'])} }}) // Lv.99 max",
        '',
        'export const LEVELS = levels as readonly LevelEntry[]',
        '',
        'const cumulativeExpByLevel: Record<number, number> = {}',
        'let cumulativeTotal = 0',
        'for (const cur of levels) {',
        '  cumulativeExpByLevel[cur.level] = cumulativeTotal',
        '  cumulativeTotal += cur.expToNext',
        '}',
        '',
        'export const CUMULATIVE_EXP_BY_LEVEL = Object.freeze(cumulativeExpByLevel)',
        '',
        'export const LEVEL_COLORS = Object.freeze(',
        '  levels.reduce<Record<number, HexColor>>((acc, cur) => {',
        '    acc[cur.level] = cur.color',
        '    return acc',
        '  }, {})',
        ')',
        '',
        'export const EXP_TO_NEXT = Object.freeze(',
        '  levels.reduce<Record<number, number>>((acc, cur) => {',
        '    acc[cur.level] = cur.expToNext',
        '    return acc',
        '  }, {})',
        ')',
        '',
        'export function getCumulativeExpForLevel(level: number): number {',
        '  const lv = Math.max(1, Math.min(99, Math.floor(level)))',
        '  return CUMULATIVE_EXP_BY_LEVEL[lv] ?? 0',
        '}',
        '',
    ])


def build_quests(quest_rows: list[dict[str, Any]]) -> str:
    rewards = [row for row in quest_rows if row.get('카테고리') == '보상']
    weights = [row for row in quest_rows if row.get('카테고리') == '생성']
    reward_lines = []
    for row in rewards:
        special = ", special: true" if row['키'] == 'Special_gold' else ''
        tier = row['키'].split('_')[0]
        if row['키'].endswith('_gold'):
            reward_lines.append(
                f"  {{ tier: {ts_string(tier)}, exp: {int(next(item['value'] for item in rewards if item['키'] == f'{tier}_exp'))}, gold: {int(row['value'])}{special} }},"
            )
    weight_map_lines = [f"  {ts_string(row['키'].replace('weight_', ''))}: {int(row['value'])}" for row in weights]
    return '\n'.join([
        "import type { QuestTierReward, QuestTier } from '../types/game'",
        '',
        'export const QUEST_REWARDS: readonly QuestTierReward[] = [',
        *reward_lines,
        '] as const',
        '',
        'export const QUEST_TIER_WEIGHTS: Readonly<Record<QuestTier, number>> = Object.freeze({',
        *[f'{line},' for line in weight_map_lines],
        '})',
        '',
    ])


def build_items(item_rows: list[dict[str, Any]]) -> str:
    materials = [row for row in item_rows if row['카테고리'] == '재료']
    others = [
        {'item_id': 'fragment_spirit_soyo', '이름': '소요의 조각'},
        {'item_id': 'fragment_spirit_rua', '이름': '루아의 조각'},
        {'item_id': 'fragment_spirit_tera', '이름': '테라의 조각'},
        {'item_id': 'fragment_spirit_pleo', '이름': '플레오의 조각'},
        {'item_id': 'fragment_spirit_stellio', '이름': '스텔리오의 조각'},
        {'item_id': 'fragment_spirit_porina', '이름': '포리나의 조각'},
        {'item_id': 'fragment_spirit_igni', '이름': '이그니의 조각'},
        {'item_id': 'fragment_spirit_nova', '이름': '노바의 조각'},
        {'item_id': 'fragment_spirit_lumen', '이름': '루멘의 조각'},
        {'item_id': 'fragment_spirit_solaris', '이름': '솔라스의 조각'},
        {'item_id': 'fragment_spirit_nubi', '이름': '누비의 조각'},
        {'item_id': 'fragment_spirit_erion', '이름': '에리온의 조각'},
        {'item_id': 'fragment_spirit_orvis', '이름': '오르비스의 조각'},
        {'item_id': 'forest_trace', '이름': '숲의 잔향'},
        {'item_id': 'wind_trace', '이름': '바람의 메아리'},
        {'item_id': 'lake_trace', '이름': '설원의 기억'},
        {'item_id': 'ruins_trace', '이름': '화염의 잔재'},
        {'item_id': 'final_trace', '이름': '어둠의 흔적'},
    ]
    english_map = {
        'flower': 'Flower',
        'leaf': 'Leaf',
        'soil': 'Soil',
        'water': 'Water',
        'fire': 'Fire',
        'wind': 'Wind',
        'star': 'Star',
        'moon': 'Moon',
        'light': 'Sun',
        'magic': 'Magic',
        'ether': 'Aether',
        'gem': 'Gem',
    }
    icon_map = {
        'fragment_spirit_soyo': 'assets/item/it/it_soul.png',
        'fragment_spirit_rua': 'assets/item/it/it_soul.png',
        'fragment_spirit_tera': 'assets/item/it/it_soul.png',
        'fragment_spirit_pleo': 'assets/item/it/it_soul.png',
        'fragment_spirit_stellio': 'assets/item/it/it_soul.png',
        'fragment_spirit_porina': 'assets/item/it/it_soul.png',
        'fragment_spirit_igni': 'assets/item/it/it_soul.png',
        'fragment_spirit_nova': 'assets/item/it/it_soul.png',
        'fragment_spirit_lumen': 'assets/item/it/it_soul.png',
        'fragment_spirit_solaris': 'assets/item/it/it_soul.png',
        'fragment_spirit_nubi': 'assets/item/it/it_soul.png',
        'fragment_spirit_erion': 'assets/item/it/it_soul.png',
        'fragment_spirit_orvis': 'assets/item/it/it_soul.png',
        'forest_trace': 'assets/item/it/it_forestmap.png',
        'wind_trace': 'assets/item/it/it_windmap.png',
        'lake_trace': 'assets/item/it/it_lakemap.png',
        'ruins_trace': 'assets/item/it/it_ruinsmap.png',
        'final_trace': 'assets/item/it/it_finalmap.png',
    }

    def item_line(row: dict[str, Any], material: bool) -> str:
        category_value = "재료" if material else "기타"
        base = [f"id: {ts_string(row['item_id'])}", f"name: {ts_string(row['이름'])}", f"category: {ts_string(category_value)}"]
        if material:
            base.append(f"englishName: {ts_string(english_map[row['item_id']])}")
            base.append(f"materialCategory: {ts_string(row['재료분류'])}")
            return '{ ' + ', '.join(base) + ' }'
        base.append(f"icon: {ts_string(icon_map[row['item_id']])}")
        return '{ ' + ', '.join(base) + ' }'

    return '\n'.join([
        "import type { CraftingMaterial, ItemDef } from '../types/game'",
        '',
        'export const CRAFTING_MATERIALS: readonly CraftingMaterial[] = [',
        *[f'\t{item_line(row, True)},' for row in materials],
        '] as const',
        '',
        'export const ETC_ITEMS: readonly ItemDef[] = [',
        *[f'\t{item_line(row, False)},' for row in others],
        '] as const',
        '',
        'export const ITEMS: readonly ItemDef[] = [...CRAFTING_MATERIALS, ...ETC_ITEMS] as const',
        '',
        'export const MATERIAL_ITEM_IDS = CRAFTING_MATERIALS.map((it) => it.id)',
        'export const ETC_ITEM_IDS = ETC_ITEMS.map((it) => it.id)',
        '',
        'export const TRACE_ITEM_BY_STAGE = {',
        "\t1: 'forest_trace',",
        "\t2: 'wind_trace',",
        "\t3: 'lake_trace',",
        "\t4: 'ruins_trace',",
        "\t5: 'final_trace',",
        '} as const',
        '',
    ])


def build_spirits(spirit_rows: list[dict[str, Any]]) -> str:
    return '\n'.join([
        "import type { SpiritDef } from '../types/game'",
        '',
        'export const SPIRITS: readonly SpiritDef[] = [',
        *[
            f"\t{{ id: {ts_string(row['spirit_id'])}, name: {ts_string(row['이름'])}, rarity: {ts_string(row['희귀도(value)'])} }},"
            for row in spirit_rows
        ],
        '] as const',
        '',
    ])


def build_recipes(recipe_rows: list[dict[str, Any]]) -> str:
    generated = [
        f"\t{{ id: {ts_string(row['recipe_id'])}, resultItemId: {ts_string(row['결과정령(value)'])}, ingredientIds: [{ts_string(row['재료1'])}, {ts_string(row['재료2'])}, {ts_string(row['재료3'])}] }},"
        for row in recipe_rows if row['결과정령(value)'] != 'TBD'
    ]
    if not generated:
        return '\n'.join([
            "import type { RecipeDef } from '../types/game'",
            '',
            '// TBD: crafting rules and failure handling',
            'export const RECIPES: readonly RecipeDef[] = [] as const',
            '',
        ])
    return '\n'.join([
        "import type { RecipeDef } from '../types/game'",
        '',
        '// TBD: crafting rules and failure handling',
        'export const RECIPES: readonly RecipeDef[] = [',
        *generated,
        '] as const',
        '',
    ])


def build_drops(exploration_rows: list[dict[str, Any]], drop_rows: list[dict[str, Any]]) -> str:
    draft_lookup = {row['키']: row['value'] for row in exploration_rows if row['카테고리'] == '공통'}
    mana_single_drop = draft_lookup.get('mana_single_drop_chance_per_expedition', 0.0001)
    return '\n'.join([
        "import type { DropDef } from '../types/game'",
        '',
        '// TBD: final region/item-specific drop rates',
        'export const DROPS: readonly DropDef[] = [] as const',
        '',
        '// Draft runtime values aligned with the exploration system spec.',
        'export const EXPEDITION_REWARD_DRAFT = {',
        f"  traceDropAmountMin: 1,",
        f"  traceDropAmountMax: 2,",
        f"  spiritFragmentDropChance: 0.72,",
        f"  spiritFragmentDropAmountMin: 1,",
        f"  spiritFragmentDropAmountMax: 2,",
        f"  manaSingleDropChancePerExpedition: {ts_string(mana_single_drop)},",
        f"  manaRewardMin: 0,",
        f"  manaRewardMax: 1,",
        f"  resultRevealDelayMs: {int(float(draft_lookup['result_reveal_delay_ms']))},",
        f"  exploreSteps: {int(float(draft_lookup['explore_steps']))},",
        f"  baseExpMin: 15,",
        f"  baseExpMax: 25,",
        f"  baseGoldMin: 10,",
        f"  baseGoldMax: 25,",
        '  eventProbabilities: {',
        f"    material: {ts_string(draft_lookup['event_material'])},",
        f"    spirit: {ts_string(draft_lookup['event_spirit'])},",
        f"    regional: {ts_string(draft_lookup['event_regional'])},",
        f"    treasure: {ts_string(draft_lookup['event_treasure'])},",
        f"    trace: {ts_string(draft_lookup['event_trace'])},",
        f"    none: {ts_string(draft_lookup['event_none'])},",
        '  },',
        '} as const',
        '',
    ])


def build_economy(economy_rows: list[dict[str, Any]]) -> str:
    tiers = [row for row in economy_rows if row['카테고리'] == '제작티어']
    tier_ranges = [(1, 19), (20, 49), (50, 79), (80, 99)]
    lookup = {row['키']: row['value'] for row in economy_rows if row.get('키')}
    return '\n'.join([
        "import type { CraftingCostTier, CraftingMaterialCost, EconomySettings } from '../types/game'",
        "import { MAIN_COLOR, TOTAL_EXP_TO_MAX } from './constants'",
        '',
        'export const ECONOMY: EconomySettings = {',
        '  mainColor: MAIN_COLOR,',
        '}',
        '',
        'export const TOTAL_EXP_REQUIRED = TOTAL_EXP_TO_MAX',
        f"export const QUEST_REJECT_PENALTY_GOLD = {int(next(row['value'] for row in economy_rows if row['키'] == 'quest_reject_penalty_gold'))} as const",
        f"export const CRAFT_SUCCESS_EXP_MIN = {int(lookup['craft_success_exp_min'])} as const",
        f"export const CRAFT_SUCCESS_EXP_MAX = {int(lookup['craft_success_exp_max'])} as const",
        f"export const CRAFT_SUCCESS_GOLD_MIN = {int(lookup['craft_success_gold_min'])} as const",
        f"export const CRAFT_SUCCESS_GOLD_MAX = {int(lookup['craft_success_gold_max'])} as const",
        f"export const CRAFT_SUCCESS_FIRST_DISCOVERY_GEM = {int(lookup['craft_success_first_discovery_gem'])} as const",
        f"export const CRAFT_FAILURE_FRAGMENT_AMOUNT = {int(lookup['craft_failure_fragment_amount'])} as const",
        f"export const CRAFT_FAILURE_EXP = {int(lookup['craft_failure_exp'])} as const",
        f"export const CRAFT_FAILURE_GOLD_MIN = {int(lookup['craft_failure_gold_min'])} as const",
        f"export const CRAFT_FAILURE_GOLD_MAX = {int(lookup['craft_failure_gold_max'])} as const",
        f"export const CRAFT_HINT_COSTS = [{int(lookup['craft_hint_gold_lv1'])}, {int(lookup['craft_hint_gold_lv2'])}, {int(lookup['craft_hint_gold_lv3'])}] as const",
        '',
        'export const LEVEL_UP_REWARD_TIERS = [',
        f"  {{ minLevel: 1, maxLevel: 19, gold: {int(lookup['levelup_gold_lv1_19'])}, mana: {int(lookup['levelup_mana_lv1_19'])} }},",
        f"  {{ minLevel: 20, maxLevel: 49, gold: {int(lookup['levelup_gold_lv20_49'])}, mana: {int(lookup['levelup_mana_lv20_49'])} }},",
        f"  {{ minLevel: 50, maxLevel: 79, gold: {int(lookup['levelup_gold_lv50_79'])}, mana: {int(lookup['levelup_mana_lv50_79'])} }},",
        f"  {{ minLevel: 80, maxLevel: 99, gold: {int(lookup['levelup_gold_lv80_99'])}, mana: {int(lookup['levelup_mana_lv80_99'])} }},",
        '] as const',
        '',
        'export function getLevelUpRewardsForLevel(level: number): { gold: number; mana: number } {',
        '  const lv = Math.max(1, Math.min(99, Math.floor(level)))',
        '  const fallback = LEVEL_UP_REWARD_TIERS[LEVEL_UP_REWARD_TIERS.length - 1]',
        '  const range = LEVEL_UP_REWARD_TIERS.find((r) => lv >= r.minLevel && lv <= r.maxLevel) ?? fallback',
        '  return { gold: range.gold, mana: range.mana }',
        '}',
        '',
        'export const CRAFTING_COST_TIERS: readonly CraftingCostTier[] = [',
        *[
            f"  {{ minLevel: {tier_ranges[index][0]}, maxLevel: {tier_ranges[index][1]}, requiredPerMaterial: {int(row['value'])}, selectedMaterialKinds: {int(next(item['value'] for item in economy_rows if item['키'] == 'selected_material_kinds'))} }},"
            for index, row in enumerate(tiers)
        ],
        '] as const',
        '',
        'export function getSpiritCraftCostByLevel(level: number): CraftingMaterialCost {',
        '  const lv = Math.max(1, Math.floor(level))',
        '  const fallback = CRAFTING_COST_TIERS[CRAFTING_COST_TIERS.length - 1]',
        '  const range = CRAFTING_COST_TIERS.find((r) => lv >= r.minLevel && lv <= r.maxLevel) ?? fallback',
        '',
        '  return {',
        '    minLevel: range.minLevel,',
        '    maxLevel: range.maxLevel,',
        '    requiredPerMaterial: range.requiredPerMaterial,',
        '    selectedMaterialKinds: range.selectedMaterialKinds,',
        '    totalMaterialCost: range.requiredPerMaterial * range.selectedMaterialKinds,',
        '  }',
        '}',
        '',
    ])


def build_regions(region_rows: list[dict[str, Any]], drop_rows: list[dict[str, Any]], current_text: str) -> str:
    templates_match = re.search(r"eventTemplates:\s*\[(.*?)\],\n\s*emptyEventTexts:", current_text, re.S)
    empty_match = re.search(r"emptyEventTexts:\s*\[(.*?)\]\n\s*  },\n\] as const", current_text, re.S)
    event_templates = templates_match.group(1).strip() if templates_match else ''
    empty_texts = empty_match.group(1).strip() if empty_match else ''
    region = {row['키']: row['value'] for row in region_rows if row.get('카테고리') == 'Map1'}
    drops = [row for row in drop_rows if row['region'] == 'starlight_forest']
    drop_lines = '\n'.join(
        f"      {{ itemId: {ts_string(row['item_id'])}, weight: {int(row['weight(value)'])}, minCount: {int(row['min_count(value)'])}, maxCount: {int(row['max_count(value)'])} }},"
        for row in drops
    )
    return '\n'.join([
        "import type { RegionDef } from '../types/game'",
        '',
        'export const REGIONS: readonly RegionDef[] = [',
        '  {',
        "    id: 'starlight_forest',",
        f"    name: {ts_string(region['region_name'])},",
        f"    unlockLevel: {int(region['unlock_level'])},",
        f"    recommendedLevel: {int(region['recommended_level'])},",
        '    manaCost: 1,',
        '    exploreSteps: 10,',
        "    traceItemId: 'forest_trace',",
        "    traceName: '숲의 잔향',",
        '    hiddenStageRequiredAmount: 20,',
        "    nextRegionId: 'wind_canyon',",
        "    nextRegionUnlockLevel: 10,",
        '    explorationRateWeights: {',
        f"      material: {int(region['weight_material'])},",
        f"      spirit: {int(region['weight_spirit'])},",
        f"      regional: {int(region['weight_regional'])},",
        f"      treasure: {int(region['weight_treasure'])},",
        '    },',
        '    discoveryTotals: {',
        f"      material: {int(region['total_material_discovery'])},",
        f"      spirit: {int(region['total_spirit_discovery'])},",
        f"      regional: {int(region['total_regional_discovery'])},",
        f"      treasure: {int(region['total_treasure_discovery'])},",
        '    },',
        '    dropTable: [',
        drop_lines,
        '    ],',
        '    eventTemplates: [',
        event_templates,
        '    ],',
        '    emptyEventTexts: [',
        empty_texts,
        '    ],',
        '  },',
        '] as const',
        '',
    ])


def main() -> int:
    workbook = load_workbook(WORKBOOK_PATH)

    player_rows = read_sheet_rows(workbook, '01_Player')
    level_rows = read_sheet_rows(workbook, '02_Level')
    exploration_rows = read_sheet_rows(workbook, '03_Exploration')
    quest_rows = read_sheet_rows(workbook, '04_Quest')
    item_rows = read_sheet_rows(workbook, '05_Item')
    drop_rows = read_sheet_rows(workbook, '06_Drop')
    economy_rows = read_sheet_rows(workbook, '09_Economy')

    files = {
        ROOT / 'src' / 'data' / 'constants.ts': build_constants(player_rows, quest_rows, economy_rows),
        ROOT / 'src' / 'data' / 'levels.ts': build_levels(level_rows),
        ROOT / 'src' / 'data' / 'quests.ts': build_quests(quest_rows),
        ROOT / 'src' / 'data' / 'items.ts': build_items(item_rows),
        ROOT / 'src' / 'data' / 'drops.ts': build_drops(exploration_rows, drop_rows),
        ROOT / 'src' / 'data' / 'economy.ts': build_economy(economy_rows),
    }

    changed: list[str] = []
    for path, content in files.items():
        if write_if_changed(path, content):
            changed.append(str(path.relative_to(ROOT)))

    if changed:
        print('updated:')
        for rel in changed:
            print(f'  - {rel}')
    else:
        print('no changes')

    return 0


if __name__ == '__main__':
    raise SystemExit(main())