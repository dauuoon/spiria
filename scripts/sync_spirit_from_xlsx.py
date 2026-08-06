from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - local setup issue
    raise SystemExit(
        'openpyxl is required. Install it with: python3 -m pip install openpyxl'
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / 'balance' / 'Spiria_Game_spirit.xlsx'
SHEET_NAME = '정령·의뢰서 합본'

NAME_TO_ID = {
    '소요': 'spirit_soyo',
    '루아': 'spirit_rua',
    '테라': 'spirit_tera',
    '플레오': 'spirit_pleo',
    '포리나': 'spirit_porina',
    '이그니': 'spirit_igni',
    '노바': 'spirit_nova',
    '루멘': 'spirit_lumen',
    '솔라스': 'spirit_solaris',
    '누비': 'spirit_nubi',
    '에리온': 'spirit_erion',
    '오르비스': 'spirit_orvis',
}

RARITY_TO_KEY = {
    '일반': 'common',
    '레어': 'rare',
    '에픽': 'epic',
    '전설': 'legendary',
}

TIER_BY_RARITY = {
    'common': 'Easy',
    'rare': 'Normal',
    'epic': 'Hard',
    'legendary': 'Special',
}

MATERIAL_NAME_TO_ID = {
    '꽃': 'flower',
    '잎': 'leaf',
    '흙': 'soil',
    '물': 'water',
    '불': 'fire',
    '바람': 'wind',
    '별': 'star',
    '달': 'moon',
    '태양': 'light',
    '빛': 'light',
    '마법': 'magic',
    '에테르': 'ether',
    '보석': 'gem',
}


def ts_string(value: Any) -> str:
    text = str(value)
    text = text.replace('\\', '\\\\').replace("'", "\\'").replace('\r\n', '\\n').replace('\n', '\\n').replace('\r', '\\n')
    return f"'{text}'"


def write_if_changed(path: Path, text: str) -> bool:
    current = path.read_text(encoding='utf-8') if path.exists() else None
    if current == text:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding='utf-8')
    return True


def parse_rows() -> list[dict[str, Any]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() if v is not None else '' for v in rows[0]]

    parsed: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}

        name = str(item.get('정령명') or '').strip()
        if not name:
            continue
        spirit_id = NAME_TO_ID.get(name)
        if not spirit_id:
            raise SystemExit(f'Unknown spirit name in workbook: {name}')

        no_value = item.get('No.')
        if not isinstance(no_value, (int, float)):
            raise SystemExit(f'Invalid discovery order for {name}: {no_value}')

        rarity_label = str(item.get('희귀도') or '').strip()
        rarity_key = RARITY_TO_KEY.get(rarity_label)
        if not rarity_key:
            raise SystemExit(f'Unknown rarity label for {name}: {rarity_label}')

        keywords_raw = str(item.get('핵심 키워드') or '').strip()
        keywords = [k.strip() for k in keywords_raw.split(',') if k and k.strip()]

        match_rate = item.get('정답 일치율')
        if isinstance(match_rate, (int, float)):
            if float(match_rate).is_integer():
                request_match_rate = f"{int(match_rate)}%"
            else:
                request_match_rate = f"{match_rate}%"
        else:
            request_match_rate = str(match_rate or '-')

        combo_raw = str(item.get('조합 순서 (1→2→3)') or '').strip()
        combo_tokens = [token.strip() for token in combo_raw.replace('->', '→').split('→') if token and token.strip()]
        if len(combo_tokens) != 3:
            raise SystemExit(f'Invalid recipe sequence for {name}: {combo_raw}')

        ingredient_ids: list[str] = []
        for token in combo_tokens:
            material_id = MATERIAL_NAME_TO_ID.get(token)
            if not material_id:
                raise SystemExit(f'Unknown material token in recipe for {name}: {token}')
            ingredient_ids.append(material_id)

        candidate_names = [
            str(item.get('정답 후보') or '').strip(),
            str(item.get('비교 후보') or '').strip(),
        ]
        candidate_ids: list[str] = []
        for candidate_name in candidate_names:
            if not candidate_name:
                continue
            candidate_id = NAME_TO_ID.get(candidate_name)
            if not candidate_id:
                raise SystemExit(f'Unknown candidate spirit name for {name}: {candidate_name}')
            if candidate_id not in candidate_ids:
                candidate_ids.append(candidate_id)
        if not candidate_ids:
            raise SystemExit(f'No candidate spirits found for request: {name}')

        parsed.append({
            'order': int(no_value),
            'id': spirit_id,
            'name': name,
            'type_label': str(item.get('유형') or '정령형').strip(),
            'rarity_label': rarity_label,
            'rarity_key': rarity_key,
            'theme_label': str(item.get('상세 배경 톤') or '따뜻한 골드').strip(),
            'story': str(item.get('정령 스토리') or '정령 소개 데이터가 준비 중입니다.').rstrip(),
            'story_box_color': str(item.get('정령 스토리 박스 색') or 'rgba(8,10,20,0.5)').strip(),
            'request_text': str(item.get('의뢰서 내용') or '의뢰서 내용이 준비 중입니다.').rstrip(),
            'keywords': keywords or ['데이터 준비중'],
            'request_match_rate': request_match_rate,
            'ingredient_ids': ingredient_ids,
            'candidate_spirit_ids': candidate_ids,
            'conversation_lines': [
                str(item.get('소통 대사 1') or f'{name}이(가) 오늘의 빛을 조심스레 건넵니다.').strip(),
                str(item.get('소통 대사 2') or f'{name}이(가) 작은 소원을 들려달라고 속삭입니다.').strip(),
                str(item.get('소통 대사 3') or f'{name}이(가) 다음 만남을 기다리며 미소 짓습니다.').strip(),
                str(item.get('소통 대사 4') or f'{name}이(가) 바람 끝에서 새로운 이야기를 들려줍니다.').strip(),
                str(item.get('소통 대사 5') or f'{name}이(가) 네 마음의 빛이 오늘도 반짝인다고 말합니다.').strip(),
                str(item.get('소통 대사 6') or f'{name}이(가) 다음 인연의 문장이 곧 이어질 거라 약속합니다.').strip(),
            ],
        })

    parsed.sort(key=lambda x: x['order'])
    return parsed


def build_spirits_ts(rows: list[dict[str, Any]]) -> str:
    spirit_lines = [
        f"\t{{ id: {ts_string(row['id'])}, name: {ts_string(row['name'])}, rarity: {ts_string(row['rarity_key'])} }},"
        for row in rows
    ]
    order_lines = [f"\t{row['id']}: {row['order']}," for row in rows]

    return '\n'.join([
        "import type { SpiritDef } from '../types/game'",
        '',
        'export const SPIRITS: readonly SpiritDef[] = [',
        *spirit_lines,
        '] as const',
        '',
        '// Discovery order from balance/Spiria_Game_spirit.xlsx (No. column).',
        'export const SPIRIT_DISCOVERY_ORDER: Readonly<Record<string, number>> = Object.freeze({',
        *order_lines,
        '})',
        '',
        'export function sortSpiritsByDiscoveryOrder(spirits: readonly SpiritDef[]): SpiritDef[] {',
        '\treturn [...spirits].sort((a, b) => {',
        '\t\tconst aOrder = SPIRIT_DISCOVERY_ORDER[a.id] ?? Number.MAX_SAFE_INTEGER',
        '\t\tconst bOrder = SPIRIT_DISCOVERY_ORDER[b.id] ?? Number.MAX_SAFE_INTEGER',
        '\t\tif (aOrder !== bOrder) return aOrder - bOrder',
        "\t\treturn a.name.localeCompare(b.name, 'ko-KR')",
        '\t})',
        '}',
        '',
        'const SPIRIT_ANIMATION_FRAMES: Readonly<Record<string, readonly [string, string, string]>> = Object.freeze({',
        "\tspirit_soyo: ['assets/spirt/soyo1.png', 'assets/spirt/soyo2.png', 'assets/spirt/soyo3.png'],",
        "\tspirit_rua: ['assets/spirt/lua1.png', 'assets/spirt/lua2.png', 'assets/spirt/lua3.png'],",
        "\tspirit_tera: ['assets/spirt/tera1.png', 'assets/spirt/tera2.png', 'assets/spirt/tera3.png'],",
        "\tspirit_pleo: ['assets/spirt/pleo1.png', 'assets/spirt/pleo2.png', 'assets/spirt/pleo3.png'],",
        "\tspirit_porina: ['assets/spirt/porina1.png', 'assets/spirt/porina2.png', 'assets/spirt/porina3.png'],",
        "\tspirit_igni: ['assets/spirt/ageuni1.png', 'assets/spirt/ageuni2.png', 'assets/spirt/ageuni3.png'],",
        "\tspirit_nova: ['assets/spirt/nova1.png', 'assets/spirt/nova2.png', 'assets/spirt/nova3.png'],",
        "\tspirit_lumen: ['assets/spirt/lumen1.png', 'assets/spirt/lumen2.png', 'assets/spirt/lumen3.png'],",
        "\tspirit_solaris: ['assets/spirt/solas1.png', 'assets/spirt/solas2.png', 'assets/spirt/solas3.png'],",
        "\tspirit_nubi: ['assets/spirt/nubi1.png', 'assets/spirt/nubi2.png', 'assets/spirt/nubi3.png'],",
        "\tspirit_erion: ['assets/spirt/erion1.png', 'assets/spirt/erion2.png', 'assets/spirt/erion3.png'],",
        "\tspirit_orvis: ['assets/spirt/orbis1.png', 'assets/spirt/orbis2.png', 'assets/spirt/orbis3.png'],",
        '})',
        '',
        'export function getSpiritAnimationFrames(spiritId: string): readonly string[] {',
        '\treturn SPIRIT_ANIMATION_FRAMES[spiritId] ?? []',
        '}',
        '',
        'export function getSpiritArtworkPath(spiritId: string): string {',
        '\tconst frames = getSpiritAnimationFrames(spiritId)',
        '\tif (frames.length > 0) return frames[0]',
        "\treturn `assets/codex/${spiritId.replace(/^spirit_/, '')}.png`",
        '}',
        '',
    ])


def build_spirit_details_ts(rows: list[dict[str, Any]]) -> str:
    blocks: list[str] = []
    for row in rows:
        keyword_items = ', '.join(ts_string(k) for k in row['keywords'])
        blocks.extend([
            f"\t{row['id']}: {{",
            f"\t\ttypeLabel: {ts_string(row['type_label'])},",
            f"\t\trarityLabel: {ts_string(row['rarity_label'])},",
            f"\t\trarityKey: {ts_string(row['rarity_key'])},",
            f"\t\tthemeLabel: {ts_string(row['theme_label'])},",
            f"\t\tstory: {ts_string(row['story'])},",
            f"\t\tstoryBoxColor: {ts_string(row['story_box_color'])},",
            f"\t\trequestText: {ts_string(row['request_text'])},",
            f"\t\tkeywords: [{keyword_items}],",
            f"\t\tconversationLines: [{', '.join(ts_string(line) for line in row['conversation_lines'])}],",
            "\t\tcraftCount: '미정',",
            f"\t\trequestMatchRate: {ts_string(row['request_match_rate'])},",
            "\t\tfirstMetDate: '미정',",
            '\t},',
        ])

    return '\n'.join([
        "import type { SpiritRarity } from '../types/game'",
        '',
        'export type SpiritThemeLabel =',
        "\t| '따뜻한 골드'",
        "\t| '차가운 푸른빛'",
        "\t| '보라빛 신비'",
        "\t| '붉은 불빛'",
        "\t| '푸릇한 그린빛'",
        '',
        'export type SpiritDetailMeta = {',
        '\ttypeLabel: string',
        '\trarityLabel: string',
        '\trarityKey: SpiritRarity',
        '\tthemeLabel: SpiritThemeLabel',
        '\tstory: string',
        '\tstoryBoxColor: string',
        '\trequestText: string',
        '\tkeywords: string[]',
        '\tconversationLines: string[]',
        '\tcraftCount: string',
        '\trequestMatchRate: string',
        '\tfirstMetDate: string',
        '}',
        '',
        'export const SPIRIT_DETAIL_META: Record<string, SpiritDetailMeta> = {',
        *blocks,
        '}',
        '',
        'export const DEFAULT_SPIRIT_DETAIL_META: SpiritDetailMeta = {',
        "\ttypeLabel: '정령형',",
        "\trarityLabel: '일반',",
        "\trarityKey: 'common',",
        "\tthemeLabel: '따뜻한 골드',",
        "\tstory: '정령 소개 데이터가 준비 중입니다.',",
        "\tstoryBoxColor: 'rgba(8,10,20,0.5)',",
        "\trequestText: '의뢰서 내용이 준비 중입니다.',",
        "\tkeywords: ['데이터 준비중'],",
        "\tconversationLines: ['안녕하세요, 오늘도 함께해요.', '조용히 반짝이는 기운이 느껴져요.', '다음에 또 이야기해요.', '오늘의 기운도 나눠 줄게요.', '네 마음의 속삭임이 들려요.', '다음 만남을 기다릴게요.'],",
        "\tcraftCount: '-',",
        "\trequestMatchRate: '-',",
        "\tfirstMetDate: '-',",
        '}',
        '',
    ])


def build_spirit_requests_ts(rows: list[dict[str, Any]]) -> str:
    blocks: list[str] = []
    for row in rows:
        tier = TIER_BY_RARITY.get(row['rarity_key'], 'Easy')
        req_id = f"req_{row['id']}"
        candidate_ids = ', '.join(ts_string(candidate_id) for candidate_id in row['candidate_spirit_ids'])
        blocks.extend([
            '  {',
            f"    id: {ts_string(req_id)},",
            f"    spiritId: {ts_string(row['id'])},",
            f"    spiritName: {ts_string(row['name'])},",
            f"    candidateSpiritIds: [{candidate_ids}],",
            f"    tier: {ts_string(tier)},",
            f"    text: {ts_string(row['request_text'])},",
            '  },',
        ])

    return '\n'.join([
        "import type { QuestTier } from '../types/game'",
        '',
        'export type SpiritRequestPage = {',
        '  id: string',
        '  spiritId: string',
        '  spiritName: string',
        '  candidateSpiritIds: string[]',
        '  candidateMatchRates?: Partial<Record<string, number>>',
        '  tier: QuestTier',
        '  text: string',
        '}',
        '',
        'export const SPIRIT_REQUEST_PAGES: readonly SpiritRequestPage[] = [',
        *blocks,
        '] as const',
        '',
    ])


def build_recipes_ts(rows: list[dict[str, Any]]) -> str:
    blocks: list[str] = []
    for row in rows:
        ingredient_ids = row['ingredient_ids']
        blocks.append(
            "  { id: %s, resultItemId: %s, ingredientIds: [%s, %s, %s] }," % (
                ts_string(f"recipe_{row['id']}"),
                ts_string(row['id']),
                ts_string(ingredient_ids[0]),
                ts_string(ingredient_ids[1]),
                ts_string(ingredient_ids[2]),
            )
        )

    return '\n'.join([
        "import type { RecipeDef } from '../types/game'",
        '',
        '// Generated from balance/Spiria_Game_spirit.xlsx (조합 순서 column).',
        'export const RECIPES: readonly RecipeDef[] = [',
        *blocks,
        '] as const',
        '',
    ])


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f'Workbook not found: {WORKBOOK_PATH}')

    rows = parse_rows()
    if not rows:
        raise SystemExit('No spirit rows were parsed from workbook')

    spirits_changed = write_if_changed(ROOT / 'src' / 'data' / 'spirits.ts', build_spirits_ts(rows))
    details_changed = write_if_changed(ROOT / 'src' / 'data' / 'spiritDetails.ts', build_spirit_details_ts(rows))
    requests_changed = write_if_changed(ROOT / 'src' / 'data' / 'spiritRequests.ts', build_spirit_requests_ts(rows))
    recipes_changed = write_if_changed(ROOT / 'src' / 'data' / 'recipes.ts', build_recipes_ts(rows))

    print(f'synced {len(rows)} spirits from {WORKBOOK_PATH.name}')
    print(f"updated src/data/spirits.ts: {'yes' if spirits_changed else 'no'}")
    print(f"updated src/data/spiritDetails.ts: {'yes' if details_changed else 'no'}")
    print(f"updated src/data/spiritRequests.ts: {'yes' if requests_changed else 'no'}")
    print(f"updated src/data/recipes.ts: {'yes' if recipes_changed else 'no'}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
